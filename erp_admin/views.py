

import datetime
from django.shortcuts import render

# Create your views here.

import os
from django.shortcuts import render

import json

from django.db import IntegrityError
from django.shortcuts import get_object_or_404, render,redirect
from django.views import View
from requests import request
from erp_admin.models import *
from erp_user.models import *
# Create your views here.
from django.contrib.admin.models import LogEntry
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.mail import BadHeaderError, send_mail
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import logout as auth_logout
from django.contrib.auth.forms import AuthenticationForm

from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.decorators import login_required

from django.contrib import messages
from django.contrib.auth.hashers import check_password, make_password
from django.views.generic.base import TemplateView
from django.contrib.auth.views import PasswordChangeView

# import pandas as pd
from django.http import JsonResponse
from django.conf import settings
from io import BytesIO
from django.http import HttpResponse

from django.utils.translation import gettext_lazy as _
# from erp_system_admin.models import CustomUser

# from erp_system_admin.models import CustomUserManager

def set_language(request):
    lang = request.GET.get('l', 'en')
    request.session[settings.LANGUAGE_SESSION_KEY] = lang
    # return to previous view
    response = HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))
    # set cookies as well
    response.set_cookie(settings.LANGUAGE_COOKIE_NAME, lang)
    return response

from django.contrib.auth import get_user_model
CustomUser = get_user_model()

def adminlogin(request):
    msg = ''
    if request.method == 'POST':
        username = request.POST['username'].strip()
        password = request.POST['password'].strip()
        role = request.POST['role'].strip()
        branch_name = request.POST['branch_name']
        branch_code = request.POST['branch_code']

        print(f"Username: {username}, Password: {password}, Role: {role}")
        user = authenticate(request, username=username, password=password)


        if user is not None:
            # Check the user's role
            if user.role == role:
                # Log the user in and create a session
                login(request, user)   
                print(f"User authenticated: {user.username}, Role: {user.role}")

                request.session['branch_name'] = branch_name
                request.session['branch_code'] = branch_code
                # Redirect based on the role
                if role == 'Admin':
                    print("Redirecting to admin dashboard")
                    return redirect('admin_dashboard')
                
                elif role == 'Employee':
                    print("Redirecting to employee dashboard")
                    return redirect('employee_dashboard')
                else:
                    messages.error(request, "Invalid role")
                    print("Invalid role provided")
            else:
                # Role doesn't match the one assigned to the user
                messages.error(request, "Role does not match the provided credentials")
                print("Role mismatch")
        else:
            # Authentication failed
            messages.error(request, "Invalid username or password")
            print("Authentication failed")
    return render(request, 'Authentication/AdminLogin.html', {'msg': msg})

def adminregister(request):
    if request.method == 'POST':
        username = request.POST['username']
        role = request.POST['role']
        hash_password = request.POST['password']
        password = make_password(hash_password)

        # Check if the username already exists
        if CustomUser.objects.filter(username=username).exists():
            messages.error(request, "Username already exists. Please choose a different username.")
            return redirect("/adminregister")
        
        # Create and save the new user
        obj = CustomUser(
            username=username,
            password=password,
            role=role,
        )
        obj.save()  
        return redirect("/adminlogin")

    return render(request, "Authentication/AdminRegister.html")

@login_required(login_url='adminlogin')
def admin_logout(request):
    logout(request)
    # Optionally, perform additional actions here
    return redirect('/adminlogin')


@login_required(login_url='adminlogin')
def admin_change_password(request):
    if 'username_session' in request.session:
        admin_username = request.session['username_session']
        admindata = admin_data.objects.get(adminusername=admin_username)

        if request.method == 'POST':
            old_password = request.POST.get('old_password')
            new_password1 = request.POST.get('new_password1')
            new_password2 = request.POST.get('new_password2')

            # Validate old password
            if not check_password(old_password, admin_data.password):
                messages.error(request, 'Incorrect current password.')
                return redirect('/admin_change_password')

            # Validate new passwords match
            if new_password1 != new_password2:
                messages.error(request, 'The new passwords do not match.')
                return redirect('/admin_change_password')

            # Update password in the admin_data object
            admin_data.password = make_password(new_password1)
            admin_data.save()

            messages.success(request, 'Your password was successfully updated!')
            return redirect('/adminlogin')  # Redirect to the admin login page or any other page
        else:
            return render(request, 'Authentication/change_password.html')
    else:
        return redirect('/adminlogin')


@login_required(login_url='adminlogin')
def admin_dashboard(request):
    # totalsales = len(salesreport.objects.all())
    totalsales = salesreport.objects.all().count()
    totalpurchase = purchasereport.objects.all().count()
    totalleadsentry = leadsentry.objects.all().count()
    totalleaves = leaves.objects.all().count()
    totaltodowork = todowork.objects.all().count()
    totalquotation = quotation_details.objects.all().count()
    return render(request, "dashboard/admin_dashboard.html",{"totalsales":totalsales,"totalpurchase":totalpurchase,"totalquotation":totalquotation,
                                                             "totalleadsentry":totalleadsentry,"totalleaves":totalleaves,"totaltodowork":totaltodowork})


# =============================================================== ALL PAGES STARTS ==============================================================



# ++++++++++++++++++++++++++++++++++++++++++++++++++++++ Branch Page Start ++++++++++++++++++++++++++++++++++++++++++++++++++

@login_required(login_url='adminlogin')
def add_branch(request):
    if request.method == "POST":
        bd = branch_data(request.POST, request.FILES)
        if bd.is_valid():
            bd.save()
            return redirect("/view_branch")
        else:
            print(bd.errors)
            messages.error(request, "The branch could not be created because the data didn't validate.")
    else:
        bd = branch_data()
    return render(request, "Branch-Page/add_branch.html", {"bd": bd})

@login_required(login_url='adminlogin')
def viewbranch(request):
    view_branchdata = branches.objects.all()
    
    return render(request,"Branch-Page/view_branch.html",{"view_branchdata":view_branchdata})


@login_required(login_url='adminlogin')
def viewbranch_fulldetails(request,branch_id):
    try:
        viewbranch_fulldetails = branches.objects.get(id=branch_id)
    except branches.DoesNotExist:
        return redirect('/view_branch')  # Redirect if branch not found
    return render(request,"Branch-Page/view_branch_fulldetails.html",{"viewbranch_fulldetails":viewbranch_fulldetails})


@login_required(login_url='adminlogin')
def deletebranch(request,branch_id):
    del_branchdata = branches.objects.get(id=branch_id).delete()
    return redirect('view_branch')
    return render(request,"Branch-Page/view_branch.html")


@login_required(login_url='adminlogin')
def updatebranch(request,branch_id):
    ubd = get_object_or_404(branches,id=branch_id)
    if request.method == "POST":
        bd = branch_data(request.POST,request.FILES,instance=ubd)
        if bd.is_valid():
            bd.save()
            return redirect('/view_branch')
        else:
            print(bd.errors)
            return render(request,"Branch-Page/add_branch.html",{"bd":bd})
    else:
        bd = branch_data(instance=ubd)
    return render(request,"Branch-Page/add_branch.html",{"bd":bd})
# ++++++++++++++++++++++++++++++++++++++++++++++++++++++ Branch Page End ++++++++++++++++++++++++++++++++++++++++++++++++++++




# ++++++++++++++++++++++++++++++++++++++++++++++++++++++ Employee reg Page Start ++++++++++++++++++++++++++++++++++++++++++++++++++

def registeremployee(request):
    if request.method == "POST":
        emp = employee_data(request.POST, request.FILES)
        username = request.POST.get('username')
        password = request.POST.get('password')

        if emp.is_valid():
                # Check if the username already exists
            if CustomUser.objects.filter(username=username).exists():
                messages.error(request, "The Username Is Already Taken. Please Choose A Different Username.")
            else:
                try:
                        # Create the User object
                    user = CustomUser.objects.create_user(username=username, password=password)
                    user.role = 'Employee'
                    user.save()

                        # Save the managerdata instance with the newly created User
                    epms = emp.save(commit=False)
                    epms.user = user  # Assign the user to the manager instance
                    epms.save()  # Save the manager instance

                    messages.success(request, "Employee Registered Successfully.")
                    return redirect("/manage_employee")
                except IntegrityError as e:
                        # Log or print the exception details for debugging
                    print(f"IntegrityError: {e}")
                    messages.error(request, "An Error Occurred While Registering The Employee. Please try again.")
        else:
            messages.error(request, "The Employee Could Not Be Registered Because The Data Didn't Validate.")
    else:
        emp = manager_data()

    view_branchdata = branches.objects.all()
    viewmanagers_details = allmanagers.objects.all()
    users = CustomUser.objects.all()

    return render(request,"Employee-Pages/register_employee.html",{"emp":emp,"viewmanagers_details":viewmanagers_details,"view_branchdata":view_branchdata})

def manageemployee(request):
    branch_name = request.session.get('branch_name')
    branch_code = request.session.get('branch_code')
    
    if branch_name and branch_code:
        try :
            branch = branches.objects.get(name=branch_name, code=branch_code)
            viewemp = allemployee.objects.filter(branch_details=branch)
        except branch.DoesNotExist:
            messages.success(request, "Branch Details Is Not Provided.")
        
    viewmanager = allmanagers.objects.all()
    return render(request,"Employee-Pages/manage_employee.html",{"viewemp":viewemp,"viewmanager":viewmanager})

def deleteemployee(request,emp_id):
    delemployee = allemployee.objects.get(id=emp_id)
    if len(delemployee.image) > 0:
        os.remove(delemployee.image.path)
    delemployee.delete()
    messages.success(request, "Employee Data Deleted Successfully.")
    return redirect('manage_employee')
    return render(request,"Employee-Pages/manage_employee.html")

def updateemployee(request,emp_id):
    uemp = get_object_or_404(allemployee,id=emp_id)
    if request.method == "POST":
        emp = employee_data(request.POST,request.FILES,instance=uemp)
        if emp.is_valid():
            emp.save()
            messages.success(request, "Employee Data Updated Successfully.")
            return redirect('/manage_employee')
        else:
            print(emp.errors)
            messages.success(request, "Employee Data Can Not Updated Successfully.")
            return render(request,"employee-Pages/register_employee.html",{"emp":emp})
    else:
        emp = employee_data(instance=uemp)
            
    viewmanagers_details = allmanagers.objects.all()
    view_branchdata = branches.objects.all()
    return render(request,"Employee-Pages/register_employee.html",{"emp":emp,"viewmanagers_details":viewmanagers_details,"view_branchdata":view_branchdata})

def employee_fulldetails(request,emp_id):
    try:
        emp_fulldetails = allemployee.objects.get(id=emp_id)
    except allemployee.DoesNotExist:
        return redirect('/view_employee')  
    
    emp_fulldetails = allemployee.objects.filter(id=emp_id).get()
    return render(request,"Employee-Pages/employee_fulldetails.html",{"emp_fulldetails":emp_fulldetails})



# ++++++++++++++++++++++++++++++++++++++++++++++++++++++ Employee reg Page End ++++++++++++++++++++++++++++++++++++++++++++++++++




# ++++++++++++++++++++++++++++++++++++++++++++++++++++++ Job role start ++++++++++++++++++++++++++++++++++++++++++++++++++

def add_jobrole(request):
    if request.method == "POST":
        addjr = jobroles(request.POST)
        if addjr.is_valid():
            addjr.save()
        else:
            print(addjr.error)
            messages.error(request,"Job Role Could Not Add")
    else:
        return redirect('add_jobrole')
    return render(request,"job_role/add_jobrole.html")

# ++++++++++++++++++++++++++++++++++++++++++++++++++++++ Job role End ++++++++++++++++++++++++++++++++++++++++++++++++++




# # ++++++++++++++++++++++++++++++++++++++++++++++ Sales Report Page Start ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@login_required(login_url='adminlogin')
def add_salesreport(request):
    if request.method == "POST":
        addsr = salesreport_data(request.POST,request.FILES)
        if addsr.is_valid():
            sales_report = addsr.save(commit=False)  # Don't save yet
            sales_report.user = request.user  # Assign the logged-in user
            sales_report.save()
            messages.success(request, "Sales Report Added Successfully")
            return redirect('/manage_salesreport')
        else:
            print(addsr.errors)
            messages.error(request,"Sales Report Can Not Add")
            return render(request,"sales_report_pages/add_sales_reports.html",{"addsr":addsr})
    else:
        addsr = salesreport_data()
        
    view_branchdata = branches.objects.all()
    return render(request,"sales_report_pages/add_sales_reports.html",{"addsr":addsr,"view_branchdata":view_branchdata})

@login_required(login_url='adminlogin')
def manage_salesreport(request):
    # view_sales_report = salesreport.objects.all()          
    # view_sales_report = salesreport.objects.all()

    # orderid_query = request.GET.get('orderid', '')
    # customer_name_query = request.GET.get('customer_name', '')
    # customer_contact_query = request.GET.get('customer_contact', '')
    # product_name_query = request.GET.get('product_name', '')
    # status_query = request.GET.get('status', '')
    # payment_method_query = request.GET.get('payment_method', '')

    
    # Filter sales report based on search inputs
    # search_data = salesreport.objects.all()
    
    # if orderid_query:
    #     view_sales_report = view_sales_report.filter(orderid__icontains=orderid_query)
    # if customer_name_query:
    #     view_sales_report = view_sales_report.filter(customer_name__icontains=customer_name_query)
    # if customer_contact_query:
    #     view_sales_report = view_sales_report.filter(customer_contact__icontains=customer_contact_query)
    # if product_name_query:
    #     view_sales_report = view_sales_report.filter(product_name__icontains=product_name_query)
    # if status_query:
    #     view_sales_report = view_sales_report.filter(status__icontains=status_query)
    # if payment_method_query:
    #     view_sales_report = view_sales_report.filter(payment_method__icontains=payment_method_query)


    search_query = request.GET.get('search_query', '').strip()
    if not search_query:
        search_data = salesreport.objects.all()

    # Filter the data
    search_data = salesreport.objects.filter(
        Q(orderid__icontains=search_query) |
        Q(dateof_sale__icontains=search_query) |
        Q(customer_name__icontains=search_query)
    )

    # If no matching records, return a message
    if not search_data.exists():
        messages.error(request,"No Matching Reocrd")
        search_data = salesreport.objects.all()
    context = {
        "search_data":search_data,
        'search_query': search_query
    }

    return render(request,"sales_report_pages/manage_sales_reports.html",context)

@login_required(login_url='adminlogin')
def delete_salesreport(request,id):
    delete_sales = salesreport.objects.get(id=id).delete()
    messages.success(request,"Sales Report Deleted Successfully")
    return redirect('/manage_salesreport')
    return render(request,"sales_report_pages/manage_sales_reports.html")

@login_required(login_url='adminlogin')
def update_salesreport(request,id):
    usales_report = get_object_or_404(salesreport,id=id)
    if request.method == "POST":
        addsr = salesreport_data(request.POST,request.FILES,instance=usales_report)
        if addsr.is_valid():
            sales_report = addsr.save(commit=False)  # Don't save yet
            sales_report.user = request.user  # Assign the logged-in user
            sales_report.save()
            messages.success(request, "Sales Report Added Successfully")
            return redirect('/manage_salesreport')
        else:
            print(addsr.errors)
            messages.error(request,"Sales Report Can Not Updated")
            return render(request,"sales_report_pages/add_sales_reports.html",{"addsr":addsr})
    else:
        addsr = salesreport_data(instance=usales_report)
        
    view_branchdata = branches.objects.all()
    return render(request,"sales_report_pages/add_sales_reports.html",{"addsr":addsr,"view_branchdata":view_branchdata})



# # ++++++++++++++++++++++++++++++++++++++++++++++ Sales Report Page End ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++




# # ++++++++++++++++++++++++++++++++++++++++++++++ Purchase Report Page Start ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@login_required(login_url='adminlogin')
def add_purchasereport(request):
    if request.method == "POST":
        apr = purchasereport_data(request.POST,request.FILES)
        if apr.is_valid():
            purchase_report = apr.save(commit=False)  # Don't save yet
            purchase_report.user = request.user  # Assign the logged-in user
            purchase_report.save()
            messages.success(request, "Purchase Report Added Successfully")
            return redirect('/manage_purchasereport')
        else:
            print(apr.errors)
            messages.error(request,"Purchase Report Can Not Add")
            return render(request,"purchase_report_pages/add_purchase_reports.html",{"apr":apr})
    else:
        apr = purchasereport_data()
        
    view_branchdata = branches.objects.all()
    return render(request,"purchase_report_pages/add_purchase_reports.html",{"apr":apr,"view_branchdata":view_branchdata})

@login_required(login_url='adminlogin')
def manage_purchasereport(request):
    # purchase_report = purchasereport.objects.all()
    purchaseid_query = request.GET.get('purchaseid')
    dateof_purchase_query = request.GET.get('dateof_purchase')
    vendor_name_query = request.GET.get('vendor_name')
    vendor_contact_query = request.GET.get('vendor_contact')
    product_name_query = request.GET.get('product_name')
    product_category_query = request.GET.get('product_category')
    quantity_purchase_query = request.GET.get('quantity_purchase')
    payment_method_query = request.GET.get('payment_method')
    unitprice_query = request.GET.get('unitprice')
    total_purchasecost_query = request.GET.get('total_purchasecost')
    status_query = request.GET.get('status')
    
    purchase_report = purchasereport.objects.all()
    
    if purchaseid_query:
        purchase_report = purchase_report.filter(purchaseid = purchaseid_query)
    if dateof_purchase_query:
        purchase_report = purchase_report.filter(dateof_purchase = dateof_purchase_query)    
    if vendor_name_query:
        purchase_report = purchase_report.filter(vendor_name = vendor_name_query)
    if vendor_contact_query:
        purchase_report = purchase_report.filter(vendor_contact = vendor_contact_query)
    if product_name_query:
        purchase_report = purchase_report.filter(product_name = product_name_query)
    if product_category_query:
        purchase_report = purchase_report.filter(product_category = product_category_query)
    if quantity_purchase_query:
        purchase_report = purchase_report.filter(quantity_purchase = quantity_purchase_query)
    if payment_method_query:
        purchase_report = purchase_report.filter(payment_method = payment_method_query)
    if unitprice_query:
        purchase_report = purchase_report.filter(unitprice = unitprice_query)
    if total_purchasecost_query:
        purchase_report = purchase_report.filter(total_purchasecost = total_purchasecost_query)
    if status_query:
        purchase_report = purchase_report.filter(status = status_query)
        
    return render(request,"purchase_report_pages/manage_purchase_reports.html",{"purchase_report":purchase_report})

@login_required(login_url='adminlogin')
def delete_purchasereport(request,id):
    delete_purchasereport = purchasereport.objects.get(id=id).delete()
    messages.success(request,"Purchase Report Deleted Successfully")
    return redirect('/manage_purchasereport')
    return render(request,"purchase_report_pages/manage_purchase_report.html")
    
@login_required(login_url='adminlogin')
def update_purchasereport(request,id):
    upurchase_report = get_object_or_404(purchasereport,id=id)
    if request.method == "POST":
        apr = purchasereport_data(request.POST,request.FILES,instance=upurchase_report)
        if apr.is_valid():
            purchase_report = apr.save(commit=False)  # Don't save yet
            purchase_report.user = request.user  # Assign the logged-in user
            purchase_report.save()
            messages.success(request, "Purchase Report Updated Successfully")
            return redirect('/manage_purchasereport')
        else:
            print(upurchase_report.errors)
            messages.error(request,"Purchase Report Can Not Updated")
            return render(request,"purchase_report_pages/add_purchase_reports.html")
    else:
        apr = purchasereport_data(instance=upurchase_report)
        
    view_branchdata = branches.objects.all()
    return render(request,"purchase_report_pages/add_purchase_reports.html",{"apr":apr,"view_branchdata":view_branchdata})


#  ++++++++++++++++++++++++++++++++++++++++++++++ Purchase Report Page End ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++


#  ++++++++++++++++++++++++++++++++++++++++++++++ leads entry page Start ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

@login_required(login_url='adminlogin')
def add_leadsentry(request):
    if request.method == "POST":
        lead = leadsentry_data(request.POST,request.FIELS)
        if lead.is_valid():
            leads = lead.save(commit=False)
            leads.user = request.user
            leads.save()
            messages.success(request,"Leads Entry Added Successfully")
            return redirect("/view_leadsentry")
        else:
            print(lead.errors)
            messages.error(request, "The Leads Entry Can Not Add Because The Data Didn't Validate.")
    else:
        lead = leadsentry_data()    
        
    view_branchdata = branches.objects.all()
    return render(request,"leads_entry_pages/add_leads_entry.html",{"lead":lead,"view_branchdata":view_branchdata})


@login_required(login_url='adminlogin')
def view_leadsentry(request):
    leads_entry = leadsentry.objects.all()
    return render(request,"leads_entry_pages/manage_leads_entry.html",{"leads_entry":leads_entry})

@login_required(login_url='adminlogin')
def update_leadsentry(request,id):
    uleads_entry = get_object_or_404(leadsentry,id=id)
    if request.method =="POST":
        lead = leadsentry_data(request.POST,request.FILES,instance=uleads_entry)
        if lead.is_valid():
            leads = lead.save(commit=False)
            leads.user = request.user
            leads.save()
            messages.success(request,"Leads Entry Updated Successfully")
            return redirect("/view_leadsentry")
        else:
            print(lead.errors)
            messages.error(request, "The Leads Entry Could Not Be Updated Because The Data Didn't Validate.")
            return redirect(request,"leads_entry_pages/add_leads_entry.html")
    else:
        lead = leadsentry_data(instance=uleads_entry)
        
    view_branchdata = branches.objects.all()
    return render(request,"leads_entry_pages/add_leads_entry.html",{"lead":lead, "view_branchdata":view_branchdata})

@login_required(login_url='adminlogin')
def delete_leadsentry(request,id):
    delete_leadsentry = leadsentry.objects.get(id=id).delete()
    messages.success(request,"Leads Entry Deleted Successfully")
    return redirect('view_leadsentry')
    return render(request,"leads_entry_pages/manage_leads_entry.html")


#  ++++++++++++++++++++++++++++++++++++++++++++++ leads entry page End ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++




#  ++++++++++++++++++++++++++++++++++++++++++++++ Quotation page Start ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

def add_quotation(request):
    if request.method == "POST":
        quotation = quotation_data(request.POST,request.FILES)
        if quotation.is_valid():
            aq = quotation.save(commit=False)
            aq.user = request.user
            aq.save()
            messages.success(request,"Quotation Added Successfully")
            return redirect('manage_quotation')
        else:
            print(quotation.errors)
            messages.error(request,"Quotation Could Not Be Added ")
    else:
        quotation = quotation_data()
        
    view_branchdata = branches.objects.all()
    return render(request,"quotation_pages/add_quotation.html",{"quotation":quotation,"view_branchdata":view_branchdata})

def manage_quotation(request):
    view_quotation = quotation_details.objects.all()
    return render(request,"quotation_pages/manage_quotation.html",{"view_quotation":view_quotation})

def delete_quotation(request,id):
    del_quotation = quotation_details.objects.get(id=id).delete()
    messages.success(request,"Quotation Deleted Successfully")
    return redirect("manage_quotation")
    return render(request,"quotation_pages/emp_view_quotation.html")

def update_quotation(request,id):
    uquotation = get_object_or_404(quotation_details,id=id)
    if request.method == "POST":
        quotation = quotation_data(request.POST,request.FILES,instance=uquotation)
        if quotation.is_valid:
            aq = quotation.save(commit=False)
            aq.user = request.user
            aq.save()
            messages.success(request,"Quotation Is Updated Successfully")
            return redirect("manage_quotation")
        else:
            print(quotation.errors)
            messages.error(request,"Quotation Could Not Updated")
    else:
        quotation = quotation_data(instance=uquotation)
    view_branchdata = branches.objects.all()
    view_emp = allemployee.objects.all()
    return render(request,"quotation_pages/add_quotation.html",{"quotation":quotation,"view_branchdata":view_branchdata,"view_emp":view_emp})
    

def quotation_fulldetails(request,id):
    view_fulldetails_quotation = quotation_details.objects.filter(id=id).get()
    return render(request,"quotation_pages/emp_quotation_fulldetails.html",{"view_fulldetails_quotation":view_fulldetails_quotation})


#  ++++++++++++++++++++++++++++++++++++++++++++++ Quotation page Start ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++




#  ++++++++++++++++++++++++++++++++++++++++++++++  to do work page Start ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

@login_required(login_url='adminlogin')
def add_todowork(request):
    if request.method == "POST":
        todo_work = todowork_data(request.POST,request.FILES)
        if todo_work.is_valid:
            todo = todo_work.save(commit=False)
            todo.user = request.user
            todo.save()
            messages.success(request,"To Do Work Added Successfully ")
            return redirect('manage_todowork')
        else:
            print(todo_work.errors)
            messages.error(request,"To Do Work Can Not Added Successfully")
    else:
        todo_work = todowork_data()
        
    allemp = allemployee.objects.all()
    view_branchdata = branches.objects.all()
    return render(request,"todo_list_pages/add_todolist.html",{"todo_work":todo_work,"allemp":allemp,"view_branchdata":view_branchdata})

@login_required(login_url='adminlogin')
def manage_todowork(request):
    todoworks = todowork.objects.all()
    return render(request,"todo_list_pages/manage_todolist.html",{"todoworks":todoworks})

@login_required(login_url='adminlogin')
def delete_todowork(request,id):
    delete_todowork = todowork.objects.get(id=id).delete()
    return redirect('manage_todowork')
    return render(request,"todo_list_pages/view_todolist.html")
    
@login_required(login_url='adminlogin')
def update_todowork(request,id):
    utodowork = get_object_or_404(todowork,id=id)
    if request.method == "POST":
        todo_work = todowork_data(request.POST,request.FILES,instance=utodowork)
        if todo_work.is_valid():
            todo = todo_work.save(commit=False)
            todo.user = request.user
            todo.save()
            messages.success(request,"To Do Work Updated Successfully")
            return redirect('manage_todowork')
        else:
            print(todo_work.errors)
            messages.error("To Do Work Can Not Updated Because Data Can Not Validate")
    else:
        todo_work = todowork_data(instance=utodowork)
    allemp = allemployee.objects.all()
    view_branchdata = branches.objects.all()
    return render(request,"todo_list_pages/add_todolist.html",{"todo_work":todo_work,'fullname': utodowork.fullname,"view_branchdata":view_branchdata,
                                                               "allemp":allemp,"status":utodowork.status,"start_date":utodowork.start_date})
#  ++++++++++++++++++++++++++++++++++++++++++++++ to do work page End ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++



# ++++++++++++++++++++++++++++++++++++++++++++++ Employee Leave Policy page Start ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

@login_required(login_url='adminlogin')
def manage_leaves(request):
    admin_viewleaves = leaves.objects.all()
    emp = allemployee.objects.all()
    view_branchdata = branches.objects.all()
    return render(request,"Leave-Page/manage_leaves.html",{"admin_viewleaves":admin_viewleaves,"emp":emp})

def admin_approve_leave(request,id):
    leave = get_object_or_404(leaves,id=id)
    leave.leavestatus = 'Approved'
    leave.save()
    return redirect('/manage_leaves')
    
def admin_disapprove_leave(request,id):
    leave = get_object_or_404(leaves,id)
    leave.leavestatus = 'Disapproved'
    leave.save()
    return redirect('/manage_leaves')
    
# @login_required(login_url='adminlogin')
# def update_leave(request,id):
#     uleave = get_object_or_404(leavedata,id)
#     if request.method == "POST":
#         leave = leavedata(request.POST,request.Files)
#     return render(request,"Leave-Page/manage_leaves.html",{"leave":leave})

# @login_required(login_url='adminlogin')

# @login_required(login_url='adminlogin')
def deleteleave(request):        
    del_leavesdata = leaves.objects.get(id=id)
    if del_leavesdata.image:
        del_leavesdata.image.delete()
    del_leavesdata.delete()
    return redirect("/manage_leaves")
    return render(request,"Admin/manage_leavepolicy_details.html")


# ++++++++++++++++++++++++++++++++++++++++++++++++ Employee Leave Policy page End ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++




# ++++++++++++++++++++++++++++++++++++++++++++++ Employee Payroll page Start ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
def addpayroll(request):
    if 'username_session' in request.session:
        if request.method == "POST":
            emp = employee_data(request.POST, request.FILES)
            if emp.is_valid():
                emp.save()
                return redirect("/manage_employee")
            else:
                print(emp.errors)
                messages.error(request, "The Employee could not be registerd because the data didn't validate.")
        else:
            emp = employee_data()
            
        viewmanagers_details = allmanagers.objects.all()
        view_branchdata = branches.objects.all()
    else:
        return redirect('/adminlogin')
    return render(request,"Employee-Pages/register_employee.html",{"emp":emp,"viewmanagers_details":viewmanagers_details,"view_branchdata":view_branchdata})

def viewpayroll(request):
    return render(request,"Payroll/add_payroll_details.html")

def deletepayroll(request):
    return render(request,"Payroll/view_payroll_details.html")

def updatepayroll(request):
    return render(request,"Payroll/view_payroll_details.html")
# ++++++++++++++++++++++++++++++++++++++++++++++++ Employee Payroll page End ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++







# ++++++++++++++++++++++++++++++++++++++++++++++ Employee Report page Start ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
def addreport(request):
    if 'username_session' in request.session:
        if request.method == "POST":
            emp = employee_data(request.POST, request.FILES)
            if emp.is_valid():
                emp.save()
                return redirect("/manage_employee")
            else:
                print(emp.errors)
                messages.error(request, "The Employee could not be registerd because the data didn't validate.")
        else:
            emp = employee_data()
            
        viewmanagers_details = allmanagers.objects.all()
        view_branchdata = branches.objects.all()
    else:
        return redirect('/adminlogin')
    return render(request,"Report-Pages/add_report.html",{"emp":emp,"viewmanagers_details":viewmanagers_details,"view_branchdata":view_branchdata})

def viewreport(request):
    return render(request,"Report-Pages/view_report_details.html")


def deletereport(request):
    return render(request,"Report-Pages/view_report_details.html")

def updatereport(request):
    return render(request,"Report-Pages/add_report_details.html")
# ++++++++++++++++++++++++++++++++++++++++++++++++ Employee Report page End ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++



# ++++++++++++++++++++++++++++++++++++++++++++++ Employee Rights And Roles page Start ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
def addrightsroles(request):
    if request.method == "POST":
        emp = employee_data(request.POST, request.FILES)
        if emp.is_valid():
            emp.save()
            return redirect("/manage_employee")
        else:
            print(emp.errors)
            messages.error(request, "The Employee could not be registerd because the data didn't validate.")
    else:
        emp = employee_data()
            
    viewmanagers_details = allmanagers.objects.all()
    view_branchdata = branches.objects.all()
    return render(request,"RightsRoles-Pages/add_rightsroles_details.html",{"emp":emp,"viewmanagers_details":viewmanagers_details,"view_branchdata":view_branchdata})

def viewrightsroles(request):
    return render(request,"RightsRoles-Pages/view_rightsroles_details.html")

def deleterightsroles(request):
    return render(request,"RightsRoles-Pages/view_rightsroles_details.html")

def updaterightsroles(request):
    return render(request,"RightsRoles-Pages/add_rightsroles_details.html")
# ++++++++++++++++++++++++++++++++++++++++++++++++ Rights And Roles page End ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++




# ++++++++++++++++++++++++++++++++++++++++++++++ Employee Work page Start ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
def addempwork(request):
    if request.method == "POST":
        emp = todowork_data(request.POST, request.FILES)
        if emp.is_valid():
            emp.save()
            return redirect("/manage_employee_work")
        else:
            print(emp.errors)
            messages.error(request, "Empwork Not Added Because Data Is Not Validate")
    else:
        emp = todowork_data()
            
    view_branchdata = branches.objects.all()
    return render(request,"Empwork-Pages/add_empwork_details.html",{"emp":emp,"view_branchdata":view_branchdata})

def viewempwork(request):
    return render(request,"Empwork-Pages/view_empwork_details.html")


def deleteempwork(request):
    return render(request,"Empwork-Pages/view_empwork_details.html")

def updateempwork(request):
    return render(request,"Empwork-Pages/add_empwork_details.html")

# ++++++++++++++++++++++++++++++++++++++++++++++++ Employee Work page End ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++




# ++++++++++++++++++++++++++++++++++++++++++++++ Attendence page Start ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

def addattendance(request, id=None):
    # Determine if this is an update or new attendance
    uatd = None
    if id:
        uatd = get_object_or_404(attendences, id=id)

    logged_in_employee = None
    logged_in_manager = None

    # Initialize form variables
    branch_details_id = None
    aemp_id = None
    am_id = None
    date = None
    total_hours = None
    status = None

    # Determine if the logged-in user is an employee or manager
    try:
        if hasattr(request.user, 'allemployee'):
            logged_in_employee = request.user.allemployee
        elif hasattr(request.user, 'allmanagers'):
            logged_in_manager = request.user.allmanagers
    except Exception as e:
        print(f"Error determining user role: {e}")

    if request.method == "POST":
        branch_details_id = request.POST.get('branch_details')
        aemp_id = request.POST.get('em')  # Employee selected
        am_id = request.POST.get('am')  # Manager selected
        date = request.POST.get('date')
        total_hours = request.POST.get('total_hours')
        status = request.POST.get('status')

        # Debugging: Print form data
        print(f"Form Data: {request.POST}")
        print(f"Branch: {branch_details_id}, Employee: {aemp_id}, Manager: {am_id}, Date: {date}, Hours: {total_hours}, Status: {status}")
            
        # Convert total_hours to a DurationField
        if total_hours:
            total_hours = timezone.timedelta(hours=float(total_hours))

        try:
            if aemp_id:
                aemp = allemployee.objects.get(id=aemp_id)
                attendance = uatd or attendences(
                    branch_details=branches.objects.get(id=branch_details_id),
                    aemp=aemp,
                    am=None,
                    date=date,
                    total_hours=total_hours,
                    status=status
                )
                attendance.save()

            elif am_id:
                am = allmanagers.objects.get(id=am_id)
                attendance = uatd or attendences(
                    branch_details=branches.objects.get(id=branch_details_id),
                    aemp=None,
                    am=am,
                    date=date,
                    total_hours=total_hours,
                    status=status
                )
                attendance.save()

            return redirect("/manage_attendance_details")
        except Exception as e:
            print(f"Error saving attendance: {e}")

    # Retrieve data for the form
    viewmanagers_details = allmanagers.objects.all()
    view_branchdata = branches.objects.all()
    view_employees = allemployee.objects.all()

    return render(request, "Attendance-Pages/add_attendance_details.html", {
        "branch_details": branch_details_id,
        "uatd": uatd,  # Pass the attendance instance for updates
        "viewmanagers_details": viewmanagers_details,
        "view_branchdata": view_branchdata,
        "view_employees": view_employees,
        "logged_in_employee": logged_in_employee,
        "logged_in_manager": logged_in_manager
    })

    
def manageattendance(request):
    emp_attendance = attendences.objects.all()
    viewmanagers_details = allmanagers.objects.all()
    view_branchdata = branches.objects.all()
    view_employees = allemployee.objects.all()
    return render(request,"Admin/manage_attendance_details.html",{"emp_attendance":emp_attendance,"viewmanagers_details":viewmanagers_details,"view_branchdata":view_branchdata,"view_employees":view_employees})


def deleteattendance(request,id):
    del_attendance = attendences.objects.get(id=id)
    del_attendance.delete()
    return render(request,"Attendance-Pages/manage_attendance_details.html")

def updateattendance(request, id):
    uatd = get_object_or_404(attendences, id=id)  # Get the attendance instance
    
    if request.method == "POST":
        ea = attendence_data(request.POST, request.FILES, instance=uatd)
        if ea.is_valid():
            ea.save()
            return redirect('/manage_attendance_details')
        else:
            print(ea.errors)
            return render(request, "Attendance-Pages/add_attendance_details.html", {
                "ea": ea,
                "uatd": uatd,  # Pass the attendance instance
                "viewmanagers_details": allmanagers.objects.all(),
                "view_employees": allemployee.objects.all(),  # Ensure employees are passed
                "view_branchdata": branches.objects.all(),
            })
    else:
        ea = attendence_data(instance=uatd)  # Pass initial form data in GET request
        viewmanagers_details = allmanagers.objects.all()
        view_branchdata = branches.objects.all()
        view_employees = allemployee.objects.all()  # Fetch all employees
        
        return render(request, "Attendance-Pages/add_attendance_details.html", {
            "ea": ea,
            "uatd": uatd,  # Pass the attendance instance for use in the template
            "viewmanagers_details": viewmanagers_details,
            "view_branchdata": view_branchdata,
            "view_employees": view_employees,  # Pass the employees list to the template
        })

# def updateattendance(request, id):
#     uatd = get_object_or_404(attendences, id=id)
    
#     if request.method == "POST":
#         ea = attendence_data(request.POST, request.FILES, instance=uatd)
#         if ea.is_valid():
#             ea.save()
#             return redirect('/manage_attendance_details')
#         else:
#             print(ea.errors)
#             return render(request, "Attendance-Pages/add_attendance_details.html", {"ea": ea})
#     else:
#         ea = attendence_data(instance=uatd)  # Pass initial form data in GET request
#         viewmanagers_details = allmanagers.objects.all()
#         view_branchdata = branches.objects.all()
#         return render(request, "Attendance-Pages/add_attendance_details.html", {
#             "ea": ea,
#             "viewmanagers_details": viewmanagers_details,
#             "view_branchdata": view_branchdata,
#         })


# def updateattendance(request,id):
#     uatd = get_object_or_404(attendences,id=id)
#     if request.method == "POST":
#         ea = attendence_data(request.POST,request.FILES,instance=uatd)
#         if ea.is_valid():
#             ea.save()
#             return redirect('/manage_attendance_details')
#         else:
#             print(ea.errors)
#             return render(request,"Attendance-Pages/add_attendance_details.html",{"ea":ea})
#     else:
#         ea = attendence_data(instance=uatd)
            
#     viewmanagers_details = allmanagers.objects.all()
#     view_branchdata = branches.objects.all()
#     return render(request,"Attendance-Pages/add_attendance_details.html",{"viewmanagers_details":viewmanagers_details,"view_branchdata":view_branchdata,"ea":ea})

# ++++++++++++++++++++++++++++++++++++++++++++++++ Attendance page End ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++



# ++++++++++++++++++++++++++++++++++++++++++++++ Data page Start ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
def adddata(request):
    if request.method == "POST":
        emp = employee_data(request.POST, request.FILES)
        if emp.is_valid():
            emp.save()
            return redirect("/manage_data")
        else:
            print(emp.errors)
            messages.error(request, "The Data Could Not Be Added because the Data Didn't Validate.")
    else:
        emp = employee_data()
            
        viewmanagers_details = allmanagers.objects.all()
        view_branchdata = branches.objects.all()
    return render(request,"Attendence-Pages/add_attendence_details.html",{"emp":emp,"viewmanagers_details":viewmanagers_details,"view_branchdata":view_branchdata})

def viewdata(request):
    return render(request,"Attendence-Pages/view_attendence_details.html")

def deletedata(request):
    return render(request,"Attendence-Pages/view_attendence_details.html")

def updatedata(request):
    return render(request,"Attendence-Pages/add_attendence_details.html")
# ++++++++++++++++++++++++++++++++++++++++++++++++ Data page End ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++


 
 
 
 
# ****************************************************************************** Employee Side ******************************************************************************

@login_required(login_url='adminlogin')
def employee_dashboard(request):
    return render(request, 'dashboard/employee_dashboard.html')

# +++++++++++++++++++++++++++++++++++++++++++++++++ Sales Report Page Start ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

@login_required(login_url='adminlogin')
def emp_add_salesreport(request):
    if request.method == "POST":
        addsr = salesreport_data(request.POST,request.FILES)
        if addsr.is_valid():
            sales_report = addsr.save(commit=False)  # Don't save yet
            sales_report.user = request.user  # Assign the logged-in user
            sales_report.save()
            messages.success(request, "Sales Report Added Successfully")
            return redirect('/employee_view_salesreport')
        else:
            print(addsr.errors)
            messages.error(request,"Sales Report Can Not Add")
            return render(request,"sales_report_pages/add_sales_reports.html",{"addsr":addsr})
    else:
        addsr = salesreport_data()
        
    view_branchdata = branches.objects.all()
    return render(request,"sales_report_pages/add_sales_reports.html",{"addsr":addsr,"view_branchdata":view_branchdata})

@login_required(login_url='adminlogin')
def emp_view_salesreport(request):
    emp_sales_report = salesreport.objects.filter(user=request.user)
    # emp_sales_report = salesreport.objects.filter(id__range=(5,20))
    return render(request,"Employee-Pages/emp_view_sales_reports.html",{"emp_sales_report":emp_sales_report})

@login_required(login_url='adminlogin')
def emp_update_salesreport(request,id):
    usales_report = get_object_or_404(salesreport,id=id)
    if request.method == "POST":
        addsr = salesreport_data(request.POST,request.FILES,instance=usales_report)
        if addsr.is_valid():
            sales_report = addsr.save(commit=False)  # Don't save yet
            sales_report.user = request.user  # Assign the logged-in user
            sales_report.save()
            messages.success(request, "Sales Report Updated Successfully")
            return redirect('/employee_view_salesreport')
        else:
            print(addsr.errors)
            messages.error(request,"Sales Report Can Not Updated")
            return render(request,"sales_report_pages/add_sales_reports.html",{"addsr":addsr})
    else:
        addsr = salesreport_data(instance=usales_report) 
        
    view_branchdata = branches.objects.all()
    return render(request,"sales_report_pages/add_sales_reports.html",{"addsr":addsr,"view_branchdata":view_branchdata})

@login_required(login_url='adminlogin')
def emp_delete_salesreport(request,id):
    delete_sales = salesreport.objects.get(id=id).delete()
    messages.success(request,"Sales Report Deleted Successfully")
    return redirect('/employee_view_salesreport')
    return render(request,"sales_report_pages/manage_sales_reports.html")


# +++++++++++++++++++++++++++++++++++++++++++++++++++ Sales Report Page End ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++


# ++++++++++++++++++++++++++++++++++++++++++++++ Purchase Report Page Start ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

@login_required(login_url='adminlogin')
def emp_add_purchasereport(request):
    if request.method == "POST":
        apr = purchasereport_data(request.POST,request.FILES)
        if apr.is_valid():
            purchase_report = apr.save(commit=False)  # Don't save yet
            purchase_report.user = request.user  # Assign the logged-in user
            purchase_report.save()
            messages.success(request,"Purchase Report Added Successfully")
            return redirect('/employee_view_purchasereport')
        else:
            print(apr.errors)
            messages.error(request,"Purchase Report Can Not Add")
            return render(request,"purchase_report_pages/add_purchase_reports.html",{"apr":apr})
    else:
        apr = purchasereport_data()
        
    view_branchdata = branches.objects.all()
    return render(request,"purchase_report_pages/add_purchase_reports.html",{"apr":apr,"view_branchdata":view_branchdata})

@login_required(login_url='adminlogin')
def emp_view_purchasereport(request):
    emp_purchase_report = purchasereport.objects.filter(user=request.user).all()
    return render(request,"Employee-Pages/emp_view_purchase_reports.html",{"emp_purchase_report":emp_purchase_report})


@login_required(login_url='adminlogin')
def emp_delete_purchasereport(request,id):
    delete_purchasereport = purchasereport.objects.get(id=id).delete()
    messages.success(request,"Purchase Report Deleted Successfully")
    return redirect('/employee_view_purchasereport')
    return render(request,"purchase_report_pages/manage_purchase_report.html")
    
@login_required(login_url='adminlogin')
def emp_update_purchasereport(request,id):
    upurchase_report = get_object_or_404(purchasereport,id=id)
    if request.method == "POST":
        apr = purchasereport_data(request.POST,request.FILES,instance=upurchase_report)
        if apr.is_valid():
            purchase_report = apr.save(commit=False)  # Don't save yet
            purchase_report.user = request.user  # Assign the logged-in user
            purchase_report.save()
            messages.success(request, "Purchase Report Updated Successfully")
            return redirect('/employee_view_purchasereport')
        else:
            print(upurchase_report.errors)
            messages.error(request,"Purchase Report Can Not Updated")
            return render(request,"purchase_report_pages/add_purchase_reports.html")
    else:
        apr = purchasereport_data(instance=upurchase_report)
    
    view_branchdata = branches.objects.all()
    return render(request,"purchase_report_pages/add_purchase_reports.html",{"apr":apr,"view_branchdata":view_branchdata})


# ++++++++++++++++++++++++++++++++++++++++++++++++ Purchase Report Page End ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

def add_vender(request):
    if request.method == "POST":
        # Bind form with POST data
        form = Vender(request.POST, request.FILES)
        if form.is_valid():
            # Save the form data to the database
            form.save()
            return redirect("/view_venders")  # Redirect to a view that lists all vendors
        else:
            # Print errors if the form is invalid
            print(form.errors)
            return render(request, "venders/add-vender.html", {"form": form})
    else:
        # If the request method is GET, render an empty form
        form = Vender()
        return render(request, "venders/add-vender.html", {"form": form})

# ++++++++++++++++++++++++++++++++++++++++++++++ To Do Work Page Start ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@login_required(login_url='adminlogin')
def emp_view_todowork(request):
    emp_todowork = todowork.objects.filter(user=request.user)
    return render(request,"Employee-Pages/emp_view_todowork.html",{"emp_todowork":emp_todowork})

@login_required(login_url='adminlogin')
def emp_add_todowork(request):
    if request.method == "POST":
        todo_work = todowork_data(request.POST,request.FILES)
        if todo_work.is_valid:
            todo = todo_work.save(commit=False)
            todo.user = request.user
            todo.save()
            messages.success(request,"To Do Work Added Successfully ")
            return redirect('employee_view_todowork')
        else:
            print(todo_work.errors)
            messages.error(request,"To Do Work Can Not Added Successfully")
    else:
        todo_work = todowork_data()
        
    allemp = allemployee.objects.filter(user=request.user)
    view_branchdata = branches.objects.all()
    return render(request,"todo_list_pages/add_todolist.html",{"todo_work":todo_work,"allemp":allemp,"view_branchdata":view_branchdata})

@login_required(login_url='adminlogin')
def emp_delete_todowork(request,id):
    delete_todowork = todowork.objects.get(id=id).delete()
    return redirect('employee_view_todowork')
    return render(request,"todo_list_pages/view_todolist.html")
    
@login_required(login_url='adminlogin')
def emp_update_todowork(request,id):
    utodowork = get_object_or_404(todowork,id=id)
    if request.method == "POST":
        todo_work = todowork_data(request.POST,request.FILES,instance=utodowork)
        if todo_work.is_valid():
            todo = todo_work.save(commit=False)
            todo.user = request.user
            todo.save()
            messages.success(request,"To Do Work Updated Successfully")
            return redirect('/employee_view_todowork')
        else:
            print(todo_work.errors)
            messages.error("To Do Work Can Not Updated Because Data Can Not Validate")
    else:
        todo_work = todowork_data(instance=utodowork)
    allemp = allemployee.objects.all()
    return render(request,"todo_list_pages/add_todolist.html",{"todo_work":todo_work,"allemp":allemp, 'status': utodowork.status,'fullname': utodowork.fullname})
# ++++++++++++++++++++++++++++++++++++++++++++++ To Do Work Page End +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++


# ++++++++++++++++++++++++++++++++++++++++++++++ Leads Entry Page Start ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
@login_required(login_url='adminlogin')
def emp_view_leadsentry(request):
    emp_leadsentry = leadsentry.objects.filter(user=request.user)
    return render(request,"Employee-Pages/emp_view_leads_entry.html",{"emp_leadsentry":emp_leadsentry})

@login_required(login_url='adminlogin')
def emp_add_leadsentry(request):
    if request.method == "POST":
        lead = leadsentry_data(request.POST,request.FILES)
        if lead.is_valid():
            leads = lead.save(commit=False)
            leads.user = request.user
            leads.save()
            messages.success(request,"Leads Entry Added Successfully")
            return redirect("/employee_view_leadsentry")
        else:
            print(lead.errors)
            messages.error(request, "The Leads Entry Can Not Add Because The Data Did Not Validate.")
    else:
        lead = leadsentry_data()    
        
    view_branchdata = branches.objects.all()
    return render(request,"leads_entry_pages/add_leads_entry.html",{"lead":lead,"view_branchdata":view_branchdata})

@login_required(login_url='adminlogin')
def emp_update_leadsentry(request,id):
    uleads_entry = get_object_or_404(leadsentry,id=id)
    if request.method =="POST":
        lead = leadsentry_data(request.POST,request.FILES,instance=uleads_entry)
        if lead.is_valid():
            leads = lead.save(commit=False)
            leads.user = request.user
            leads.save()
            messages.success(request,"Leads Entry Updated Successfully")
            return redirect("/employee_view_leadsentry")
        else:
            print(lead.errors)
            messages.error(request, "The Leads Entry Could Not Be Updated Because The Data Didn't Validate.")
            return redirect(request,"leads_entry_pages/add_leads_entry.html")
    else:
        lead = leadsentry_data(instance=uleads_entry)
        
    view_branchdata = branches.objects.all()
    return render(request,"leads_entry_pages/add_leads_entry.html",{"lead":lead,"view_branchdata":view_branchdata})

@login_required(login_url='adminlogin')
def emp_delete_leadsentry(request,id):
    delete_leadsentry = leadsentry.objects.get(id=id).delete()
    messages.success(request,"Leads Entry Deleted Successfully")
    return redirect('employee_view_leadsentry')
    return render(request,"leads_entry_pages/manage_leads_entry.html")
# ++++++++++++++++++++++++++++++++++++++++++++++ Leads Entry Page End +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++


# +++++++++++++++++++++++++++++++++++++++++++++++ Apply For Leave page Start ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
def employee_applyfor_leave(request):
    if request.method == "POST":
        employee_id = request.POST.get('employee')
        alemp = leavedata(request.POST, request.FILES)
        if alemp.is_valid():
            leave = alemp.save(commit=False)
            leave.user = request.user
            leave.employee = alemp.cleaned_data['employee']
            leave.save()
            messages.success(request,"Leave Added Successfully")
            return redirect("/employee_view_leave")
        else:
            print(alemp.errors)
            messages.error(request, "The Leave Could Not Be Added Because Data Did Not Validate")
    else:
        alemp = leavedata()
    
    view_branchdata = branches.objects.all()
    view_emp = allemployee.objects.filter(user=request.user)
    return render(request,"Employee-Pages/apply_for_leave.html",{"view_branchdata":view_branchdata,"view_emp":view_emp,"alemp":alemp})

def employee_viewleave(request):
    view_leavesdata = leaves.objects.filter(user=request.user)
    return render(request,"Employee-Pages/emp_view_leaves.html",{"view_leavesdata":view_leavesdata})
    
def employee_deleteleave(request):     
    del_leavesdata = leaves.objects.get(id=id)
    if del_leavesdata !=[]:
        del_leavesdata.delete()
        messages.success(request,"Leave Deleted Successfully")
        return redirect("/employee_view_leave")
    return render(request,"Employee-Pages/emp_view_leaves.html")

def employee_updateleave(request,id):
    ueleave = get_object_or_404(leaves,id=id)
    if request.method == "POST":
        alemp = leavedata(request.POST,request.FILES,instance=ueleave)
        if alemp.is_valid():
            leave = alemp.save(commit=False)
            leave.user = request.user
            leave.save()
            messages.success(request,"Leave Updated Successfully")
            return redirect('/employee_view_leave')
        else:
            print(leave.errors)
            messages.error(request,"Leave Can Not Update Because Data Did Not Validate ")
            return render(request,"Employee-Pages/emp_view_leaves.html",{"ueleave":ueleave})
    else:
        alemp = leavedata(instance=ueleave)
    view_branchdata = branches.objects.all()
    view_emp = allemployee.objects.all()
    
    return render(request,"Employee-Pages/apply_for_leave.html",{"view_branchdata":view_branchdata,"alemp":alemp,"view_emp":view_emp})


# ++++++++++++++++++++++++++++++++++++++++++++++++ Apply For Leave page End ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++





# ++++++++++++++++++++++++++++++++++++++++++++++++ Quotation Page Start ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
def emp_addquotation(request):
    if request.method == "POST":
        quotation = quotation_data(request.POST,request.FILES)
        if quotation.is_valid():
            aq = quotation.save(commit=False)
            aq.user = request.user
            aq.save()
            messages.success(request,"Quotation Added Successfully")
            return redirect('employee_view_quotation')
        else:
            print(quotation.errors)
            messages.error(request,"Quotation Could Not Be Added ")
    else:
        quotation = quotation_data()
        
    view_branchdata = branches.objects.all()
    return render(request,"quotation_pages/add_quotation.html",{"quotation":quotation,"view_branchdata":view_branchdata})

def emp_viewquotation(request):
    view_quotation = quotation_details.objects.filter(user=request.user)
    return render(request,"Employee-Pages/emp_view_quotation.html",{"view_quotation":view_quotation})

def emp_quotation_fulldetails(request,id):
    view_fulldetails_quotation = quotation_details.objects.filter(id=id).get()
    return render(request,"quotation_pages/emp_quotation_fulldetails.html",{"view_fulldetails_quotation":view_fulldetails_quotation})

def emp_deletequotation(request,id):
    del_quotation = quotation_details.objects.get(id=id).delete()
    messages.success(request,"Quotation Deleted Successfully")
    return redirect("employee_view_quotation")
    return render(request,"Employee-Pages/emp_view_quotation.html")

def emp_updatequotation(request,id):
    uquotation = get_object_or_404(quotation_details,id=id)
    if request.method == "POST":
        quotation = quotation_data(request.POST,request.FILES,instance=uquotation)
        if quotation.is_valid:
            aq = quotation.save(commit=False)
            aq.user = request.user
            aq.save()
            messages.success(request,"Quotation Is Updated Successfully")
            return redirect("employee_view_quotation")
        else:
            print(quotation.errors)
            messages.error(request,"Quotation Could Not Updated")
    else:
        quotation = leavedata(instance=uquotation)
    view_branchdata = branches.objects.all()
    view_emp = allemployee.objects.filter(user=request.user)
    return render(request,"quotation_pages/add_quotation.html",{"quotation":quotation,"view_branchdata":    view_branchdata,"view_emp":view_emp})


# ++++++++++++++++++++++++++++++++++++++++++++++++ Quotation page End ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++




# ******************************************************************* Employee Side End ******************************************************************








# ++++++++++++++++++++++++++++++++++++++++++++++++++++++ Manager Page Start ++++++++++++++++++++++++++++++++++++++++++++++++++



# @login_required(login_url='adminlogin')
# def registermanager(request):
#     if request.method == "POST":
#         regm = manager_data(request.POST, request.FILES)
#         username = request.POST.get('username')
#         password = request.POST.get('password')

#         if regm.is_valid():
#             # Check if the username already exists
#             if CustomUser.objects.filter(username=username).exists():
#                 messages.error(request, "The Username Is Already Taken. Please Choose A Different Username.")
#             else:
#                 try:
#                     user = CustomUser.objects.create_user(username=username, password=password)
#                     user.role = 'Manager'
#                     user.save()
#                     manager = regm.save(commit=False)
#                     manager.user = user
#                     manager.save()

#                     messages.success(request, "Manager Registered Successfully.")
#                     return redirect("/manage_managers")
#                 except IntegrityError as e:
#                     print(f"IntegrityError: {e}")
#                     messages.error(request, "An error occurred while registering the manager. Please try again.")
#         else:
#             messages.error(request, "The Manager Could Not Be Registered Because The Data Didn't Validate.")
#     else:
#         regm = manager_data()

#     view_branchdata = branches.objects.all()
#     users = CustomUser.objects.all()

#     return render(request, "Manager-Pages/register_manager.html", {
#         "regm": regm,
#         "view_branchdata": view_branchdata,
#         "users": users
#     })


# def managemanager(request):
#     viewmanagers_details = allmanagers.objects.all()
#     return render(request,"Manager-Pages/manage_manager.html",{"viewmanagers_details":viewmanagers_details})

# def deletemanager(request,manager_id):
#     delmanager = allmanagers.objects.get(id=manager_id)
#     if len(delmanager.image) > 0:
#         os.remove(delmanager.image.path)
#     delmanager.delete()
#     return redirect('manage_managers')
#     return render(request,"Manager-Pages/manage_manager.html")

# def updatemanager(request,manager_id):
#     uregm = get_object_or_404(allmanagers,id=manager_id)
#     if request.method == "POST":
#         regm = manager_data(request.POST,request.FILES,instance=uregm)
#         if regm.is_valid():
#             regm.save()
#             return redirect('/manage_managers')
#         else:
#             print(regm.errors)
#             return render(request,"Manager-Pages/register_manager.html",{"regm":regm})
#     else:
#         regm = manager_data(instance=uregm)
#         view_branchdata = branches.objects.all()
#     return render(request,"Manager-Pages/register_manager.html",{"regm":regm, "view_branchdata": view_branchdata})

# # def manager_fulldetails(request,manager_id):
# #     managerfulldetails = managers.objects.filter(id=manager_id).get()
# #     return render(request,"manager_fulldetails.html",{"managerfulldetails":managerfulldetails})

# def manager_fulldetails(request,manager_id):
#     try:
#         viewmanager_fulldetails = allmanagers.objects.get(id=manager_id)
#     except allmanagers.DoesNotExist:
#         return redirect('/view_manager')  
    
#     return render(request,"Manager-Pages/manager_fulldetails.html",{"viewmanager_fulldetails":viewmanager_fulldetails})

 
# ++++++++++++++++++++++++++++++++++++++++++++++++++++++ Manager Page End ++++++++++++++++++++++++++++++++++++++++++++++++++

# ============================================================================================================== ALL PAGES END ==============================================================================================






from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import os
from django.http import HttpResponse


#  ======================== generate pdf ====================================

# def render_to_pdf(template_src,context_dict={}):
#     template = get_template(template_src)
#     html = template.render(context_dict)
#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Desposition'] = 'attechment; filename="report.pdf"'
#     pisa_status = pisa.CreatePDF(html, dest=response)

#     if pisa_status.err:
#         return HttpResponse('Error In Generating Pdf', status=404)
#     return response


def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html  = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None

def download_salesreport_pdf(request,sales_id):
    template = get_template('pdfs/invoice.html')
    sales_report = get_object_or_404(salesreport, id=sales_id)
    context = {'sales_report':sales_report}
    html = template.render(context)
    pdf = render_to_pdf('pdfs/invoice.html', context)
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = f"Invoice_{sales_id}.pdf"
        content = f"inline; filename='{filename}'"
        download = request.GET.get("download")
        if download:
            content = f"attachment; filename='{filename}'"
        response['Content-Disposition'] = content
        return response
    return HttpResponse("Not found",status=404) 

#  ======================== download pdf ====================================

# def download_pdf(request,sales_id):
#     sales_report = get_object_or_404(salesreport, id=sales_id)
    
#     if not sales_report:
#         return HttpResponse("No sales report found.", status=404)
#     context = {'sales_report' : sales_report}
#     pdf = render_to_pdf('sales_report_pages/manage_sales_reports.html',context)
    
#     return pdf


# @login_required(login_url='adminlogin')
# def download_excel(request):
#     view_sales_report = salesreport.objects.all()      
#     data = []
#     for obj in view_sales_report:
#         data.append({
#             "salesid":obj.orderid,
#             "dateof_sale":obj.dateof_sale,
#             "customer_name":obj.customer_name,
#             "customer_contact":obj.customer_contact,
#             "product_name":obj.product_name,
#             "product_category":obj.product_category,
#             "quantity_sold":obj.quantity_sold,
#             "unitprice":obj.unitprice,
#             "total":obj.total,
#             "payment_methods":obj.payment_methods,
#             "status":obj.status
#         })
        
#         output = BytesIO()
#         pd.DataFrame(data).to_excel('sales_reports.xlsx',index=False)
#         output.seek(0)
        
#         response = HttpResponse(output,content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Desposition'] = 'attechment; filename=sales_reports.xlsx'
#         messages.success(request,'Excel File Download Successfully')
#         return redirect('manage_salesreport')
#     messages.error(request,'Excel File Could Not Be Downloaded')
#     return render(request,"sales_report_pages/manage_sales_reports.html",{"view_sales_report":view_sales_report})


import xlwt
from django.http import HttpResponse
from django.db.models import Q
import openpyxl
import logging

logger = logging.getLogger(__name__)

def export_sales_to_excel(request):
    search_query = request.GET.get('search_query', '').strip()
    
    if not search_query:
        return HttpResponse("No search query provided.")

    # Filter the data
    search_data = salesreport.objects.filter(
        Q(orderid__icontains=search_query) |
        Q(dateof_sale__icontains=search_query) |
        Q(customer_name__icontains=search_query)
    )

    # If no matching records, return a message
    if not search_data.exists():
        return HttpResponse("No matching records found for the search query.")
    
        # Create an Excel file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Filtered Sales Report"
    
    headers = ["orderid","Date of Sales","Customer Name"]
    sheet.append(headers)
    
    for my_row in search_data:
        sheet.append([my_row.orderid,
                    my_row.dateof_sale.strftime('%d-%m-%Y') if my_row.dateof_sale else '',
                    my_row.customer_name])
        
    response = HttpResponse(content_type="application/ms-excel")
        
    response['Content-Disposition'] = 'attachment; filename=filtered_data' + str(datetime.datetime.now()) +'.xlsx'
    workbook.save(response)

    return response
    return render(request, 'sales_report_pages/manage_sales_reports.html', {'search_data': search_data, 'search_query': search_query})

    

def download_excel_salesreport(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="excelsalesreports.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet("sheet1")
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    colums = ['orderid','dateof_sale','customer_name','customer_contact','product_name','product_category','quantity_sold','payment_method','unitprice','total','status']
    for col_num in range(len(colums)):
        ws.write(row_num, col_num,colums[col_num],font_style)
    font_style = xlwt.XFStyle()
    
    # Get individual search parameters from the request
    orderid_query = request.GET.get('orderid', '')
    customer_name_query = request.GET.get('customer_name', '')
    customer_contact_query = request.GET.get('customer_contact', '')
    product_name_query = request.GET.get('product_name', '')
    status_query = request.GET.get('status', '')
    payment_method_query = request.GET.get('payment_method', '')

    # Filter the data using Q objects and combining the conditions
    data = salesreport.objects.all()

    if orderid_query:
        data = data.filter(orderid__icontains=orderid_query)
    if customer_name_query:
        data = data.filter(customer_name__icontains=customer_name_query)
    if customer_contact_query:
        data = data.filter(customer_contact__icontains=customer_contact_query)
    if product_name_query:
        data = data.filter(product_name__icontains=product_name_query)
    if status_query:
        data = data.filter(status__icontains=status_query)
    if payment_method_query:
        data = data.filter(payment_method__icontains=payment_method_query)
    
    for my_row in data:
        row_num += 1
        ws.write(row_num, 0, my_row.orderid, font_style)
        ws.write(row_num, 1, my_row.dateof_sale, font_style)
        ws.write(row_num, 2, my_row.customer_name, font_style)
        ws.write(row_num, 3, my_row.customer_contact, font_style)
        ws.write(row_num, 4, my_row.product_name, font_style)
        ws.write(row_num, 5, my_row.product_category, font_style)
        ws.write(row_num, 6, my_row.quantity_sold, font_style)
        ws.write(row_num, 7, my_row.payment_method, font_style)
        ws.write(row_num, 8, my_row.unitprice, font_style)
        ws.write(row_num, 9, my_row.total, font_style)
        ws.write(row_num, 10, my_row.status, font_style)
    wb.save(response)
    return response


def download_excel_purchasereport(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="excel_purchasereports.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet("sheet1")
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    columns = ['purchaseid','dateof_purchase','vendor_name','vendor_contact','product_name','product_category','quantity_purchase','payment_method',
              'unitprice','total_purchasecost','status']
    for col_num in range(len(columns)):
        ws.write(row_num,col_num,columns[col_num], font_style)
    font_style = xlwt.XFStyle()
    
    purchaseid_query = request.GET.get('purchaseid', '')
    dateof_purchase_query = request.GET.get('dateof_purchase', '')
    vendor_name_query = request.GET.get('vendor_name', '')
    vendor_contact_query = request.GET.get('vendor_contact', '')
    product_name_query = request.GET.get('product_name', '')
    product_category_query = request.GET.get('product_category', '')
    quantity_purchase = request.GET.get('product_category', '')
    payment_method = request.GET.get('product_category', '')
    unitprice = request.GET.get('product_category', '')
    total_purchasecost = request.GET.get('product_category', '')
    status = request.GET.get('product_category', '')
    
    data = purchasereport.objects.all()
    
    if purchaseid_query:
        data = data.filter(purchaseid__icontains=purchaseid_query)
    if dateof_purchase_query:
        data = data.filter(dateof_purchase__icontains=dateof_purchase_query)   
    if vendor_name_query:
        data = data.filter(vendor_name__icontains=vendor_name_query)
    if vendor_contact_query:
        data = data.filter(vendor_contact__icontains=vendor_contact_query)
    if product_name_query:
        data = data.filter(product_name__icontains=product_name_query)
    if product_category_query:
        data = data.filter(product_category__icontains=product_category_query)
    
    for my_row in data:
        row_num += 1
        ws.write(row_num, 0, my_row.purchaseid, font_style)
        ws.write(row_num, 1, my_row.dateof_purchase, font_style)
        ws.write(row_num, 2, my_row.vendor_name, font_style)
        ws.write(row_num, 3, my_row.vendor_contact, font_style)
        ws.write(row_num, 4, my_row.product_name, font_style)
        ws.write(row_num, 5, my_row.product_category, font_style)
        ws.write(row_num, 6, my_row.quantity_purchase, font_style)
        ws.write(row_num, 7, my_row.payment_method, font_style)
        ws.write(row_num, 8, my_row.unitprice, font_style)
        ws.write(row_num, 9, my_row.total_purchasecost, font_style)
        ws.write(row_num, 10, my_row.status, font_style)
    wb.save(response)
    return response


def download_excel_leadsentry(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="excel_leadsentry.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet("sheet1")
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    columns = ['fullname','email','phone','companyname','jobtitle','website','source','followup_date','followup_method','budget','leadstatus','leadowner']
    for col_num in range(len(columns)):
        ws.write(row_num,col_num,columns[col_num], font_style)
    font_style = xlwt.XFStyle()
    data = leadsentry.objects.all()
    for my_row in data:
        row_num += 1
        ws.write(row_num, 0, my_row.fullname, font_style)
        ws.write(row_num, 1, my_row.email, font_style)
        ws.write(row_num, 2, my_row.phone, font_style)
        ws.write(row_num, 3, my_row.companyname, font_style)
        ws.write(row_num, 4, my_row.jobtitle, font_style)
        ws.write(row_num, 5, my_row.website, font_style)
        ws.write(row_num, 6, my_row.source, font_style)
        ws.write(row_num, 7, my_row.followup_date, font_style)
        ws.write(row_num, 8, my_row.followup_method, font_style)
        ws.write(row_num, 9, my_row.budget, font_style)
        ws.write(row_num, 10, my_row.leadstatus, font_style)
        ws.write(row_num, 11, my_row.leadowner, font_style)
    wb.save(response)
    return response


def download_excel_leaves(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="excel_leaves.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet("sheet1")
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    columns = ['branch_details','employee','image','gender','leavesdate','leaveedate','nofdays','reasontype','reason','document','leavestatus','email','mobile']
    for col_num in range(len(columns)):
        ws.write(row_num,col_num,columns[col_num], font_style)
    font_style = xlwt.XFStyle()
    data = leaves.objects.all()
    for my_row in data:
        row_num += 1
        ws.write(row_num, 0, my_row.branch_details, font_style)
        ws.write(row_num, 1, my_row.employee, font_style)
        ws.write(row_num, 2, my_row.image, font_style)
        ws.write(row_num, 3, my_row.gender, font_style)
        ws.write(row_num, 4, my_row.leavesdate, font_style)
        ws.write(row_num, 5, my_row.leaveedate, font_style)
        ws.write(row_num, 6, my_row.nofdays, font_style)
        ws.write(row_num, 7, my_row.reasontype, font_style)
        ws.write(row_num, 8, my_row.reason, font_style)
        ws.write(row_num, 9, my_row.document, font_style)
        ws.write(row_num, 10, my_row.leavestatus, font_style)
        ws.write(row_num, 11, my_row.email, font_style)
        ws.write(row_num, 12, my_row.mobile, font_style)
    wb.save(response)
    return response

def download_excel_todowork(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="excel_todowork.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet("sheet1")
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    columns = ['fullname','work','start_date','end_date','work_desc','website','status','delay_reason']
    for col_num in range(len(columns)):
        ws.write(row_num,col_num,columns[col_num], font_style)
    font_style = xlwt.XFStyle()
    data = todowork.objects.all()
    for my_row in data:
        row_num += 1
        ws.write(row_num, 0, my_row.fullname, font_style)
        ws.write(row_num, 1, my_row.work, font_style)
        ws.write(row_num, 2, my_row.start_date, font_style)
        ws.write(row_num, 3, my_row.end_date, font_style)
        ws.write(row_num, 4, my_row.work_desc, font_style)
        ws.write(row_num, 5, my_row.status, font_style)
        ws.write(row_num, 6, my_row.delay_reason, font_style)
    wb.save(response)
    return response

def download_excel_data(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="excel_salesreports.xls"'

	#creating workbook
	wb = xlwt.Workbook(encoding='utf-8')
	
	ws = wb.add_sheet("sheet1")
	row_num = 0
	font_style = xlwt.XFStyle()
	font_style.font.bold = True
	columns = ['ID', 'Consignee Name', 'Consignee Address', 'PI NO', 'Date','Clearing Port','Product Name','Exchange Rate','Fortyfive Container',
                'Forty Container','twenty_container','sl_qty_cont','sl_inr_per_cont','sl_gst_18','sl_total_inr','cfs_qty_cont','cfs_inr_per_cont',
                'cfs_gst_18','cfs_total_inr','transportation_qty_cont','transportation_inr_per_cont','transportation_gst_18','transportation_total_inr',
                'stamp_qty_cont','stamp_inr_per_cont','stamp_gst_18','stamp_total_inr','agency_qty_cont','agency_inr_per_cont','agency_gst_18',
                'agency_total_inr','customduty_qty_cont','customduty_inr_per_cont','customduty_gst_18','customduty_total_inr','oceanfreight_qty_cont',
                'oceanfreight_inr_per_cont','oceanfreight_gst_18','oceanfreight_total_inr','round_off','total','ac_name','bank_name','ac_no','ifsc_code',
                'pan_no','gst_no',
            ]

	for col_num in range(len(columns)):
		ws.write(row_num, col_num, columns[col_num], font_style)
	font_style = xlwt.XFStyle()
	data = quotation_details.objects.all() 
	for my_row in data:
		row_num = row_num + 1
		ws.write(row_num, 0, my_row.id, font_style)
		ws.write(row_num, 1, my_row.consignee_name, font_style)
		ws.write(row_num, 2, my_row.consignee_address, font_style)
		ws.write(row_num, 3, my_row.pi_no, font_style)
	wb.save(response)
	return response
# def download_pdf(request):
#     sales_report = salesreport.objects.filter(user=request.user)
#     context = {'sales_report' : sales_report}
#     html = get_template("sales_report_pages/manage_sales_reports.html").render(context)
#     pdf_file = HTML(string = html).write_pdf()
    
#     response = HttpResponse(pdf_file,content_type='application/pdf')
#     response['Content-Desposition'] = 'attechment; filename="report.pdf"'
#     return response



#  ======================== page not found ====================================
def error_404(request, exception):
    data = {"name": "somthing error"}
    return render(request,'404.html', data)




